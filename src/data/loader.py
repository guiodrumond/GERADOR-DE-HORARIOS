from pathlib import Path
from openpyxl import load_workbook
import logging

from src.domain.database import BaseDados
from src.domain.models import (
    Curso,
    Turma,
    Especialidade,
    ParPedagogico,
    PadraoPedagogico,
    Slot,
    Restricao,
    Atribuicao,
)

class ExcelLoader:
    def __init__(self, caminho_arquivo: str):
        caminho = Path(caminho_arquivo)
        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        logging.info(f"Carregando dados do arquivo: {caminho.name}")
        self.wb = load_workbook(filename=caminho, data_only=True)

    # ==========================================================
    # UTILITÁRIOS
    # ==========================================================
    def _sheet(self, nome: str):
        if nome not in self.wb.sheetnames:
            raise ValueError(f"Aba não encontrada no Excel: '{nome}'. Verifique o arquivo.")
        return self.wb[nome]

    def _linhas(self, aba: str):
        ws = self._sheet(aba)
        linhas = list(ws.iter_rows(values_only=True))

        if not linhas:
            return []

        # Limpa os cabeçalhos para evitar erros com espaços (ex: "Curso " -> "Curso")
        cabecalho = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(linhas[0])]
        registros = []

        for linha in linhas[1:]:
            # Pula linhas inteiramente em branco
            if all(c is None or str(c).strip() == "" for c in linha):
                continue
            registros.append(dict(zip(cabecalho, linha)))

        return registros

    def _texto(self, valor):
        return str(valor).strip() if valor is not None else ""

    def _inteiro(self, valor):
        try:
            return int(valor) if valor is not None else 0
        except ValueError:
            return 0

    def _booleano(self, valor):
        texto = self._texto(valor).upper()
        return texto in {"S", "SIM", "TRUE", "1", "V", "VERDADEIRO"}

    # ==========================================================
    # CARREGADORES ESPECÍFICOS (Usando .get para segurança)
    # ==========================================================
    def load_config(self):
        config = {}
        for row in self._linhas("CONFIG"):
            parametro = self._texto(row.get("Parâmetro"))
            if parametro:
                config[parametro] = row.get("Valor")
        return config

    def load_cursos(self):
        cursos = []
        for row in self._linhas("CURSOS"):
            cursos.append(
                Curso(
                    nome_curso=self._texto(row.get("Nome_Curso")),
                    codigo=self._texto(row.get("Curso")),
                    especialidade_ftp=self._texto(row.get("Especialidade_FTP")),
                    padrao_ftp=self._texto(row.get("Padrao_FTP")),
                )
            )
        return cursos

    def load_turmas(self):
        turmas = []
        for row in self._linhas("TURMAS"):
            turmas.append(
                Turma(
                    codigo=self._texto(row.get("Turma")),
                    curso=self._texto(row.get("Curso")),
                    padrao_ftp=self._texto(row.get("Padrao_FTP")),
                    ativa=self._booleano(row.get("Ativa")),
                )
            )
        return turmas

    def load_especialidades(self):
        especialidades = []
        for row in self._linhas("ESPECIALIDADES_1ANO"):
            especialidades.append(
                Especialidade(
                    id=self._inteiro(row.get("Id")),
                    componente=self._texto(row.get("Componente")),
                    nome=self._texto(row.get("Especialidade")),
                    sigla=self._texto(row.get("Sigla")),
                    aulas=self._inteiro(row.get("Aulas")),
                )
            )
        return especialidades

    def load_pares_pedagogicos(self):
        pares = []
        for row in self._linhas("PARES_PEDAGOGICOS"):
            pares.append(
                ParPedagogico(
                    codigo=self._texto(row.get("Par")),
                    especialidade_1=self._texto(row.get("Especialidade_1")),
                    especialidade_2=self._texto(row.get("Especialidade_2")),
                )
            )
        return pares

    def load_padroes_pedagogicos(self):
        padroes = []
        for row in self._linhas("PADROES_PEDAGOGICOS"):
            padroes.append(
                PadraoPedagogico(
                    componente=self._texto(row.get("Componente")),
                    tipo=self._texto(row.get("Tipo")),
                    valor=self._texto(row.get("Valor")),
                )
            )
        return padroes

    def load_slots(self):
        slots = []
        for row in self._linhas("SLOTS"):
            slots.append(
                Slot(
                    dia=self._texto(row.get("Dia")),
                    aula=self._inteiro(row.get("Aula")),
                )
            )
        return slots

    def load_restricoes(self):
        restricoes = []
        for row in self._linhas("RESTRICOES"):
            restricoes.append(
                Restricao(
                    regra=self._texto(row.get("Regra")),
                    valor=self._texto(row.get("Valor")),
                )
            )
        return restricoes

    def load_atribuicoes(self):
        atribuicoes = []
        for row in self._linhas("ATRIBUICOES"):
            atribuicoes.append(
                Atribuicao(
                    turma=self._texto(row.get("Turma")),
                    especialidade=self._texto(row.get("Especialidade")),
                    professor=self._texto(row.get("Professor")),
                )
            )
        return atribuicoes

    # ==========================================================
    # BASE DADOS (Com liberação de memória)
    # ==========================================================
    def load(self) -> BaseDados:
        try:
            return BaseDados(
                config=self.load_config(),
                cursos=self.load_cursos(),
                turmas=self.load_turmas(),
                especialidades=self.load_especialidades(),
                pares_pedagogicos=self.load_pares_pedagogicos(),
                padroes_pedagogicos=self.load_padroes_pedagogicos(),
                slots=self.load_slots(),
                restricoes=self.load_restricoes(),
                atribuicoes=self.load_atribuicoes(),
            )
        finally:
            # Garante que a planilha será fechada da memória mesmo se houver erro no parse
            self.wb.close()