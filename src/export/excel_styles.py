from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ExcelStyles:

    TITLE_FILL = "34495E"
    HEADER_FILL = "34495E"
    EMPTY_FILL = "F2F3F4"
    BORDER_COLOR = "DDDDDD"

    AREA_COLORS = {
        "MEST": {"bg": "D6EAF8", "font": "2874A6"},
        "CNST": {"bg": "D5F5E3", "font": "1E8449"},
        "CHSA": {"bg": "FADBD8", "font": "943126"},
        "LEST": {"bg": "FFE6CC", "font": "935116"},
        "FTP": {"bg": "E8DAEF", "font": "6C3483"},
        "PV": {"bg": "D1F2EB", "font": "117A65"},
        "PA": {"bg": "E5E7E9", "font": "5D6D7A"},
        "PROJ": {"bg": "E5E7E9", "font": "5D6D7A"},
        "PLANEJAMENTO": {"bg": "C6E0B4", "font": "385723"},
        "DEFAULT": {"bg": "FFFFFF", "font": "000000"}
    }

    @classmethod
    def thin_border(cls):

        side = Side(
            style="thin",
            color=cls.BORDER_COLOR,
        )

        return Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )

    @classmethod
    def title_fill(cls):

        return PatternFill(
            fill_type="solid",
            fgColor=cls.TITLE_FILL,
        )

    @classmethod
    def header_fill(cls):

        return PatternFill(
            fill_type="solid",
            fgColor=cls.HEADER_FILL,
        )

    @classmethod
    def empty_fill(cls):

        return PatternFill(
            fill_type="solid",
            fgColor=cls.EMPTY_FILL,
        )

    @classmethod
    def area_fill(
        cls,
        area,
    ):

        if not area:

            cor = cls.EMPTY_FILL

        else:

            cor = cls.AREA_COLORS.get(
                area.upper(),
                cls.AREA_COLORS["DEFAULT"],
            )["bg"]

        return PatternFill(
            fill_type="solid",
            fgColor=cor,
        )

    @classmethod
    def apply_title(
        cls,
        cell,
    ):

        cell.font = Font(
            bold=True,
            color="FFFFFF",
            size=14,
        )

        cell.fill = cls.title_fill()

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    @classmethod
    def apply_header(
        cls,
        cell,
    ):

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = cls.header_fill()

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.border = cls.thin_border()

    @classmethod
    def apply_body(
        cls,
        cell,
        texto,
        area=None,
    ):

        font_color = "000000"
        is_bold = False

        if texto:
            if area:
                estilo = cls.AREA_COLORS.get(
                    area.upper(), 
                    cls.AREA_COLORS["DEFAULT"]
                )
                font_color = estilo["font"]
                is_bold = True

            cell.fill = cls.area_fill(area)
        else:
            cell.fill = cls.empty_fill()

        cell.font = Font(
            bold=is_bold,
            color=font_color,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = cls.thin_border()

    @classmethod
    def apply_aula_cell(
        cls,
        cell,
    ):

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = cls.header_fill()

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.border = cls.thin_border()