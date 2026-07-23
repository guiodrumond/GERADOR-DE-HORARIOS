from pathlib import Path
from openpyxl import Workbook
from src.export.excel_styles import ExcelStyles
from src.export.dashboard_exporter import DashboardExporter

class ExcelExporter:
    def __init__(self, base, solver=None, tempo_decorrido=0.0, analise_areas=None):
        self.base = base
        self.solver = solver
        self.tempo_decorrido = tempo_decorrido
        self.analise_areas = analise_areas
        
        # Mapas auxiliares para cruzar siglas, áreas e professores
        self.componente_para_grupo = self._criar_mapa_componente_grupo()
        self.professores_por_area = self._criar_mapa_professores_por_area()

    def export(self, grid, caminho_saida: str):
        caminho = Path(caminho_saida)
        caminho.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        aba_padrao = workbook.active
        workbook.remove(aba_padrao)

        # 1. Cria a Aba de Painel / Resumo na primeira posição (Índice 0)
        sheet_dashboard = workbook.create_sheet(title="RESUMO", index=0)
        dashboard = DashboardExporter(base=self.base, solver=self.solver, tempo_decorrido=self.tempo_decorrido, analise_areas=self.analise_areas)
        dashboard.export(sheet_dashboard)

        # 2. Cria as demais abas de horários
        sheet_turmas = workbook.create_sheet(title="TURMAS")
        self._write_turmas_sheet(sheet=sheet_turmas, grid=grid)

        sheet_professores = workbook.create_sheet(title="PROFESSORES")
        self._write_professores_sheet(sheet=sheet_professores, grid=grid)

        sheet_reunioes = workbook.create_sheet(title="REUNIÕES COLETIVAS")
        self._write_reunioes_sheet(sheet=sheet_reunioes, grid=grid)

        workbook.save(caminho)
        return caminho

    def _write_turmas_sheet(self, sheet, grid):
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        current_row = 1
        for turma in self._turmas_ativas():
            current_row = self._write_turma_table(
                sheet=sheet,
                grid=grid,
                turma=turma.codigo,
                dias=dias,
                aulas=aulas,
                start_row=current_row,
            )
            current_row += 2

    def _write_professores_sheet(self, sheet, grid):
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        professores = self._professores_do_grid(grid)
        current_row = 1

        for professor in professores:
            current_row = self._write_professor_table(
                sheet=sheet,
                grid=grid,
                professor=professor,
                dias=dias,
                aulas=aulas,
                start_row=current_row,
            )
            current_row += 2

    def _write_turma_table(self, sheet, grid, turma, dias, aulas, start_row):
        total_colunas = len(dias) + 1

        sheet.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas
        )
        
        title_cell = sheet.cell(row=start_row, column=1, value=f"HORÁRIO {turma}")
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 28

        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2

        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)
            sheet.row_dimensions[row].height = 45

            for col_offset, dia in enumerate(dias, start=2):
                cell_data = grid.get(turma, dia, aula)
                texto = ""
                grupo = None

                if cell_data:
                    componente = cell_data.texto
                    professor = cell_data.professor
                    grupo = self.componente_para_grupo.get(componente.upper())

                    if professor:
                        texto = f"{componente}\n{professor}"
                    else:
                        texto = componente

                cell = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell, texto, grupo)

        self._merge_blocos_turma(
            sheet=sheet,
            grid=grid,
            turma=turma,
            dias=dias,
            aulas=aulas,
            body_start=body_start,
        )

        return body_start + len(aulas)

    def _write_professor_table(self, sheet, grid, professor, dias, aulas, start_row):
        total_colunas = len(dias) + 1

        sheet.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas
        )
        
        title_cell = sheet.cell(row=start_row, column=1, value=professor)
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 28

        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2

        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)
            sheet.row_dimensions[row].height = 45

            for col_offset, dia in enumerate(dias, start=2):
                # Desempacota as 3 variáveis retornadas pelo novo método
                entradas, is_planning, bloco_id = self._entradas_professor_slot(
                    grid=grid, professor=professor, dia=dia, aula=aula
                )

                if is_planning:
                    texto = "\n".join(list(dict.fromkeys(entradas))) 
                    grupo = "PLANEJAMENTO"
                elif entradas:
                    texto = "\n".join(entradas)
                    
                    # MAGIA AQUI: Decide a cor com base no bloco verdadeiro, ignorando o texto!
                    if bloco_id == "ATIVIDADE":
                        grupo = "ATIVIDADE"
                    elif bloco_id == "A_DEFINIR":
                        grupo = "A DEFINIR"
                    elif bloco_id == "PROJETO":
                        grupo = texto.upper()
                    else:
                        componente = entradas[0].split(" - ")[-1].strip()
                        grupo = self.componente_para_grupo.get(componente.upper())
                else:
                    texto = ""
                    grupo = None

                cell = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell, texto, grupo)

        return body_start + len(aulas)

    def _write_header(self, sheet, dias, row):
        cell = sheet.cell(row=row, column=1, value="AULA")
        ExcelStyles.apply_header(cell)

        for coluna, dia in enumerate(dias, start=2):
            cell = sheet.cell(row=row, column=coluna, value=dia)
            ExcelStyles.apply_header(cell)

    def _merge_blocos_turma(self, sheet, grid, turma, dias, aulas, body_start):
        for col_offset, dia in enumerate(dias, start=2):
            inicio = None
            chave_atual = None

            for index, aula in enumerate(aulas):
                row = body_start + index
                cell_data = grid.get(turma, dia, aula)

                chave_merge = (
                    (cell_data.bloco_id, cell_data.texto, cell_data.professor)
                    if cell_data else None
                )

                if chave_merge is not None and chave_merge == chave_atual:
                    continue

                if chave_atual is not None and inicio is not None and row - inicio > 1:
                    self._merge_range(sheet, inicio, row - 1, col_offset)

                inicio = row
                chave_atual = chave_merge

            ultima_row = body_start + len(aulas) - 1
            if chave_atual is not None and inicio is not None and ultima_row - inicio >= 1:
                self._merge_range(sheet, inicio, ultima_row, col_offset)

    def _merge_range(self, sheet, start_row, end_row, column):
        sheet.merge_cells(
            start_row=start_row, start_column=column, end_row=end_row, end_column=column
        )

    def _entradas_professor_slot(self, grid, professor, dia, aula):
        entradas = []
        is_planning = False
        bloco_encontrado = None
        
        for turma in grid.turmas():
            cell = grid.get(turma, dia, aula)
            if not cell or cell.professor != professor:
                continue

            bloco_encontrado = cell.bloco_id
                
            if cell.bloco_id == "PLANEJAMENTO":
                entradas.append(cell.texto)
                is_planning = True
            elif cell.bloco_id in ["PROJETO", "ATIVIDADE", "A_DEFINIR"]:
                entradas.append(cell.texto) # Imprime limpo, sem o ID da turma virtual
            else:
                entradas.append(f"{turma} - {cell.texto}")
                
        return entradas, is_planning, bloco_encontrado

    def _professores_do_grid(self, grid):
        professores = set()
        for turma in grid.turmas():
            for dia in self._dias():
                for aula in self._aulas():
                    cell = grid.get(turma, dia, aula)
                    if cell and cell.professor:
                        professores.add(cell.professor)
        return sorted(professores)

    def _setup_general_sheet(self, sheet):
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 10
        
        dias = self._dias()
        for col in range(2, len(dias) + 2):
            # Blindagem: Usa get_column_letter para evitar erro em células mescladas
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = 25

    def _turmas_ativas(self):
        return [turma for turma in self.base.turmas if turma.ativa]

    def _dias(self):
        dias = []
        for slot in self.base.slots:
            if slot.dia not in dias:
                dias.append(slot.dia)
        return dias

    def _aulas(self):
        return sorted({slot.aula for slot in self.base.slots})

    def _criar_mapa_componente_grupo(self):
        mapa = {}
        for esp in self.base.especialidades:
            mapa[esp.sigla.upper()] = esp.componente.upper()
        return mapa

    def _criar_mapa_professores_por_area(self):
        mapa = {}
        for atribuicao in self.base.atribuicoes:
            if not atribuicao.professor:
                continue
            
            sigla = atribuicao.especialidade.upper()
            area = self.componente_para_grupo.get(sigla)
            
            if area:
                if area not in mapa:
                    mapa[area] = set()
                mapa[area].add(atribuicao.professor)
                
        return {area: list(profs) for area, profs in mapa.items()}
    
    def _write_reunioes_sheet(self, sheet, grid):
        """Constrói o quadro geral focando apenas no nome/tipo da reunião, ignorando totalmente os professores."""
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        start_row = 1
        total_colunas = len(dias) + 1

        # Título da aba
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas)
        title_cell = sheet.cell(row=start_row, column=1, value="QUADRO GERAL DE PLANEJAMENTOS / REUNIÕES COLETIVAS")
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 30

        # Cabeçalho dos dias da semana
        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2
        siglas_conhecidas = ["LEST", "CNST", "CHSA", "FTP", "MEST", "ART", "POR", "EDF", "ING"]

        # Preenchendo as aulas
        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)

            max_grupos_neste_horario = 1

            for col_offset, dia in enumerate(dias, start=2):
                reunioes_encontradas = set()
                grupo_estilo = "PLANEJAMENTO"
                
                for turma in grid.turmas():
                    cell = grid.get(turma, dia, aula)
                    # Verifica se é um bloco de planejamento
                    if cell and cell.bloco_id == "PLANEJAMENTO":
                        # Pega o atributo que define o nome da reunião ou o bloco, ignorando o nome do professor
                        nome_reuniao = getattr(cell, 'nome_reuniao', None) or getattr(cell, 'subtipo', None) or "PLANEJAMENTO COLETIVO"
                        
                        # Se o texto da célula contiver o nome de um professor, tratamos para exibir o tipo da reunião
                        if cell.texto and cell.texto in reunioes_encontradas:
                            continue
                            
                        # Vamos extrair limpo: se o texto não for o nome de um professor, usamos ele; senão, usamos o bloco_id
                        texto_candidato = cell.texto if cell.texto and not any(prof in cell.texto.upper() for prof in ["ALEX", "ALINE", "ARTHUR", "CAIQUE", "CLARA"]) else "PLAN ÁREA"
                        
                        if nome_reuniao not in reunioes_encontradas:
                            reunioes_encontradas.add(nome_reuniao)
                            
                        # Detecção de cor inteligente baseada nas siglas
                        texto_upper = nome_reuniao.upper()
                        for sigla in siglas_conhecidas:
                            if sigla in texto_upper:
                                grupo_estilo = self.componente_para_grupo.get(sigla, sigla)
                                break

                if reunioes_encontradas:
                    texto = "\n".join(sorted(list(reunioes_encontradas)))
                    max_grupos_neste_horario = max(max_grupos_neste_horario, len(reunioes_encontradas))
                else:
                    texto = ""

                cell_excel = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell_excel, texto, grupo_estilo)

            sheet.row_dimensions[row].height = max(45, max_grupos_neste_horario * 20)
        """Constrói o quadro geral de reuniões coletivas com detecção inteligente de cores por palavras-chave."""
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        start_row = 1
        total_colunas = len(dias) + 1

        # Título da aba
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas)
        title_cell = sheet.cell(row=start_row, column=1, value="QUADRO GERAL DE PLANEJAMENTOS / REUNIÕES COLETIVAS")
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 30

        # Cabeçalho dos dias da semana
        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2

        # Lista de palavras-chave conhecidas para mapear cores automaticamente
        # (Você pode adicionar mais termos aqui se criarem novas áreas)
        siglas_conhecidas = ["LEST", "CNST", "CHSA", "FTP", "MEST", "ART", "POR", "EDF", "ING"]

        # Preenchendo as aulas
        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)

            max_grupos_neste_horario = 1

            for col_offset, dia in enumerate(dias, start=2):
                grupos_em_reuniao = []
                grupo_estilo = "PLANEJAMENTO" # Cor genérica padrão se não achar lógica
                
                for turma in grid.turmas():
                    cell = grid.get(turma, dia, aula)
                    if cell and cell.bloco_id == "PLANEJAMENTO" and cell.texto:
                        texto_original = cell.texto.strip()
                        if texto_original not in grupos_em_reuniao:
                            grupos_em_reuniao.append(texto_original)
                        
                        # --- DETETIVE INTELIGENTE DE CORES ---
                        texto_upper = texto_original.upper()
                        
                        # Varre o texto procurando se alguma sigla conhecida (ex: CHSA, LEST) está lá dentro
                        sigla_encontrada = None
                        for sigla in siglas_conhecidas:
                            if sigla in texto_upper:
                                sigla_encontrada = sigla
                                break
                        
                        if sigla_encontrada:
                            # Se achou uma sigla conhecida, tenta usar o estilo dela
                            grupo_estilo = self.componente_para_grupo.get(sigla_encontrada, sigla_encontrada)
                        else:
                            # Se for um nome sem lógica (ex: "REUNIÃO ARTICULAÇÃO"), 
                            # tenta ver se o dicionário da base reconhece o texto inteiro ou deixa no padrão
                            grupo_estilo = self.componente_para_grupo.get(texto_upper, "PLANEJAMENTO")

                if grupos_em_reuniao:
                    texto = "\n".join(sorted(grupos_em_reuniao))
                    max_grupos_neste_horario = max(max_grupos_neste_horario, len(grupos_em_reuniao))
                else:
                    texto = ""

                # Pinta a célula com o resultado do detetive
                cell_excel = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell_excel, texto, grupo_estilo)

            sheet.row_dimensions[row].height = max(45, max_grupos_neste_horario * 20)
        """Constrói um quadro geral listando quais grupos de planejamento estão ocorrendo em cada slot."""
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        start_row = 1
        total_colunas = len(dias) + 1

        # Título da aba
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas)
        title_cell = sheet.cell(row=start_row, column=1, value="QUADRO GERAL DE PLANEJAMENTOS / REUNIÕES COLETIVAS")
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 30

        # Cabeçalho dos dias da semana
        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2

        # Preenchendo as aulas
        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)

            max_grupos_neste_horario = 1

            for col_offset, dia in enumerate(dias, start=2):
                grupos_em_reuniao = set()
                
                # A MÁGICA MUDA AQUI: 
                # Em vez de pegar o 'cell.professor', nós pegamos o 'cell.texto' (Ex: PLAN LEST)
                for turma in grid.turmas():
                    cell = grid.get(turma, dia, aula)
                    if cell and cell.bloco_id == "PLANEJAMENTO" and cell.texto:
                        grupos_em_reuniao.add(cell.texto)

                # Prepara o texto da célula de forma limpa
                if grupos_em_reuniao:
                    lista_grupos = sorted(list(grupos_em_reuniao))
                    texto = "\n".join(lista_grupos)
                    max_grupos_neste_horario = max(max_grupos_neste_horario, len(lista_grupos))
                else:
                    texto = ""

                # Pinta a célula 
                cell_excel = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell_excel, texto, "PLANEJAMENTO")

            # Altura fixa e agradável, parecida com o grid padrão
            sheet.row_dimensions[row].height = max(45, max_grupos_neste_horario * 20)
        """Constrói um quadro geral listando quais professores estão em planejamento em cada slot."""
        dias = self._dias()
        aulas = self._aulas()
        self._setup_general_sheet(sheet)

        start_row = 1
        total_colunas = len(dias) + 1

        # Título da aba
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_colunas)
        title_cell = sheet.cell(row=start_row, column=1, value="QUADRO GERAL DE PLANEJAMENTOS / REUNIÕES COLETIVAS")
        ExcelStyles.apply_title(title_cell)
        sheet.row_dimensions[start_row].height = 30

        # Cabeçalho dos dias da semana
        header_row = start_row + 1
        self._write_header(sheet=sheet, dias=dias, row=header_row)

        body_start = start_row + 2

        # Preenchendo as aulas
        for row_offset, aula in enumerate(aulas):
            row = body_start + row_offset
            
            aula_cell = sheet.cell(row=row, column=1, value=aula)
            ExcelStyles.apply_aula_cell(aula_cell)

            max_profs_neste_horario = 1 # Usado para esticar a altura da linha dinamicamente

            for col_offset, dia in enumerate(dias, start=2):
                profs_em_reuniao = set()
                
                # Vasculha a escola inteira procurando quem está em "PLANEJAMENTO" nesse dia e aula
                for turma in grid.turmas():
                    cell = grid.get(turma, dia, aula)
                    if cell and cell.bloco_id == "PLANEJAMENTO" and cell.professor:
                        profs_em_reuniao.add(cell.professor)

                # Prepara o texto da célula
                if profs_em_reuniao:
                    lista_profs = sorted(list(profs_em_reuniao))
                    texto = "\n".join(lista_profs)
                    max_profs_neste_horario = max(max_profs_neste_horario, len(lista_profs))
                else:
                    texto = ""

                # Pinta a célula (aproveitamos a formatação do ExcelStyles para o grupo "PLANEJAMENTO")
                cell_excel = sheet.cell(row=row, column=col_offset, value=texto)
                ExcelStyles.apply_body(cell_excel, texto, "PLANEJAMENTO")

            # Estica a altura da linha baseada em quantos professores caíram juntos (ex: 15 pontos de altura por professor)
            sheet.row_dimensions[row].height = max(45, max_profs_neste_horario * 16)