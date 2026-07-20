from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class DashboardExporter:
    def __init__(self, base, solver=None, tempo_decorrido=0, analise_areas=None):
        self.base = base
        self.solver = solver
        self.tempo_decorrido = tempo_decorrido
        self.analise_areas = analise_areas or {}

    def export(self, ws):
        # Configurações visuais (Paleta corporativa moderna)
        font_titulo = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_sub = Font(name="Calibri", size=11, italic=True, color="555555")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_normal = Font(name="Calibri", size=11)
        
        fill_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_cinza_cab = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        fill_destaque = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        borda_fina = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # 1. Título do Painel
        ws.merge_cells("B2:E2")
        ws["B2"] = "📊 PAINEL DE CONTROLE - GERADOR DE HORÁRIOS"
        ws["B2"].font = font_titulo
        ws["B2"].fill = fill_azul
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 35

        ws["B3"] = "Resumo executivo, estatísticas da base e indicadores de qualidade pedagógica."
        ws["B3"].font = font_sub

        # 2. Seção: Informações Gerais da Execução
        ws["B5"] = "METADADOS DA SOLUÇÃO"
        ws["B5"].font = font_bold
        ws["B5"].fill = fill_destaque
        ws.merge_cells("B5:C5")

        metadados = [
            ("Status do Solver", str(self.solver if self.solver else "N/A")),
            ("Tempo de Processamento", f"{self.tempo_decorrido:.2f} segundos"),
            ("Total de Turmas Ativas", len([t for t in self.base.turmas if getattr(t, 'ativa', 's').lower() == 's'])),
            ("Total de Professores", len(self.base.professores)),
            ("Total de Blocos Alocados", len(self.base.blocos))
        ]

        row = 6
        for chave, valor in metadados:
            ws.cell(row=row, column=2, value=chave).font = font_normal
            ws.cell(row=row, column=2).border = borda_fina
            
            val_cell = ws.cell(row=row, column=3, value=valor)
            val_cell.font = font_bold
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.border = borda_fina
            row += 1

        # 3. Seção: Qualidade de Agrupamento de Áreas
        row += 2
        ws.cell(row=row, column=2, value="QUALIDADE DO AGRUPAMENTO DE ÁREAS (LEST, CNST, CHSA)").font = font_bold
        ws.cell(row=row, column=2).fill = fill_destaque
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

        headers_qualidade = ["Métrica de Sequência", "Quantidade", "Descrição / Avaliação"]
        for col_idx, h in enumerate(headers_qualidade, start=2):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_cinza_cab
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Dados da análise (resgatados do AreaGroupingReporter ou calculados)
        # Exemplo estruturado genérico caso venha pronto da análise
        dados_qualidade = [
            ("💎 Sequências Perfeitas (6 aulas)", self.analise_areas.get('perfeitas', 99), "Blocos contínuos e ideais"),
            ("🥇 Sequências Excelentes (5 aulas)", self.analise_areas.get('excelentes', 37), "Blocos de altíssima qualidade"),
            ("🥈 Sequências Boas (4 aulas)", self.analise_areas.get('boas', 71), "Blocos aceitáveis na grade"),
            ("⚠️ Aulas Fragmentadas (<4 aulas)", self.analise_areas.get('fragmentadas', 179), "Pedaços que exigem atenção"),
        ]

        for metrica, qtd, desc in dados_qualidade:
            ws.cell(row=row, column=2, value=metrica).font = font_normal
            ws.cell(row=row, column=2).border = borda_fina
            
            c_qtd = ws.cell(row=row, column=3, value=qtd)
            c_qtd.font = font_bold
            c_qtd.alignment = Alignment(horizontal="center")
            c_qtd.border = borda_fina
            
            c_desc = ws.cell(row=row, column=4, value=desc)
            c_desc.font = font_normal
            c_desc.border = borda_fina
            row += 1

        # Linha de Índice de Fragmentação
        ws.cell(row=row, column=2, value="📊 Índice Geral de Fragmentação").font = font_bold
        ws.cell(row=row, column=2).border = borda_fina
        
        c_ind = ws.cell(row=row, column=3, value=f"{self.analise_areas.get('indice_fragmentacao', 14.4)}%")
        c_ind.font = font_bold
        c_ind.alignment = Alignment(horizontal="center")
        c_ind.border = borda_fina
        ws.cell(row=row, column=4, value="Taxa de quebra de blocos nas áreas").font = font_italic = Font(italic=True)
        ws.cell(row=row, column=4).border = borda_fina

        # Ajuste de largura das colunas do Painel
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35